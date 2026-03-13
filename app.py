"""
Video OCR Extractor v3.0
- 사용자별 추출 이력 저장 (SQLite)
- CSV / Excel / PDF 파일 이력 목록
- 탭 구조: 추출 / 이력 / 사용자 관리
"""

import streamlit as st
import cv2
import numpy as np
import pytesseract
import pandas as pd
import tempfile
import os, io, time, json, re, hashlib, sqlite3
from datetime import datetime
from PIL import Image
from streamlit_drawable_canvas import st_canvas
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from fpdf import FPDF

# ══════════════════════════════════════════════════════════
# PAGE CONFIG
# ══════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Video OCR Extractor v3.0",
    page_icon="🎬",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════════
# DATABASE  (SQLite — persists while container is alive)
# ══════════════════════════════════════════════════════════
DB_PATH = os.path.join(os.path.dirname(__file__), "ocr_history.db")

def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT    UNIQUE NOT NULL,
            created  TEXT    DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS sessions (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER NOT NULL,
            video_name  TEXT,
            interval_s  REAL,
            ocr_mode    TEXT,
            lang        TEXT,
            region      TEXT,
            total_rows  INTEGER DEFAULT 0,
            avg_conf    REAL    DEFAULT 0,
            created     TEXT    DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS rows (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER NOT NULL,
            timecode   TEXT,
            time_sec   REAL,
            confidence INTEGER,
            value      TEXT,
            FOREIGN KEY(session_id) REFERENCES sessions(id)
        );
        CREATE TABLE IF NOT EXISTS exports (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER NOT NULL,
            user_id    INTEGER NOT NULL,
            fmt        TEXT,
            filename   TEXT,
            size_bytes INTEGER,
            created    TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(session_id) REFERENCES sessions(id)
        );
        """)

init_db()

def upsert_user(username: str) -> int:
    with get_db() as conn:
        conn.execute("INSERT OR IGNORE INTO users (username) VALUES (?)", (username,))
        row = conn.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
        return row["id"]

def save_session(user_id, video_name, interval_s, ocr_mode, lang, region, results) -> int:
    avg_conf = sum(r["신뢰도(%)"] for r in results) / len(results) if results else 0
    region_str = json.dumps(region) if region else ""
    with get_db() as conn:
        cur = conn.execute(
            """INSERT INTO sessions (user_id,video_name,interval_s,ocr_mode,lang,region,total_rows,avg_conf)
               VALUES (?,?,?,?,?,?,?,?)""",
            (user_id, video_name, interval_s, ocr_mode, lang, region_str, len(results), round(avg_conf, 1))
        )
        sid = cur.lastrowid
        conn.executemany(
            "INSERT INTO rows (session_id,timecode,time_sec,confidence,value) VALUES (?,?,?,?,?)",
            [(sid, r["타임코드"], r["시간(초)"], r["신뢰도(%)"], r["추출값"]) for r in results]
        )
        return sid

def log_export(session_id, user_id, fmt, filename, size_bytes):
    with get_db() as conn:
        conn.execute(
            "INSERT INTO exports (session_id,user_id,fmt,filename,size_bytes) VALUES (?,?,?,?,?)",
            (session_id, user_id, fmt, filename, size_bytes)
        )

def get_all_users():
    with get_db() as conn:
        return conn.execute("SELECT * FROM users ORDER BY created DESC").fetchall()

def get_sessions(user_id=None):
    with get_db() as conn:
        if user_id:
            return conn.execute(
                "SELECT s.*,u.username FROM sessions s JOIN users u ON s.user_id=u.id WHERE s.user_id=? ORDER BY s.created DESC",
                (user_id,)
            ).fetchall()
        return conn.execute(
            "SELECT s.*,u.username FROM sessions s JOIN users u ON s.user_id=u.id ORDER BY s.created DESC"
        ).fetchall()

def get_rows(session_id):
    with get_db() as conn:
        return conn.execute("SELECT * FROM rows WHERE session_id=? ORDER BY time_sec", (session_id,)).fetchall()

def get_exports(user_id=None):
    with get_db() as conn:
        if user_id:
            return conn.execute(
                """SELECT e.*,u.username,s.video_name FROM exports e
                   JOIN users u ON e.user_id=u.id
                   JOIN sessions s ON e.session_id=s.id
                   WHERE e.user_id=? ORDER BY e.created DESC""", (user_id,)
            ).fetchall()
        return conn.execute(
            """SELECT e.*,u.username,s.video_name FROM exports e
               JOIN users u ON e.user_id=u.id
               JOIN sessions s ON e.session_id=s.id
               ORDER BY e.created DESC"""
        ).fetchall()

def delete_session(session_id):
    with get_db() as conn:
        conn.execute("DELETE FROM rows WHERE session_id=?", (session_id,))
        conn.execute("DELETE FROM exports WHERE session_id=?", (session_id,))
        conn.execute("DELETE FROM sessions WHERE id=?", (session_id,))

def delete_user(user_id):
    with get_db() as conn:
        sids = [r[0] for r in conn.execute("SELECT id FROM sessions WHERE user_id=?", (user_id,)).fetchall()]
        for sid in sids:
            conn.execute("DELETE FROM rows WHERE session_id=?", (sid,))
            conn.execute("DELETE FROM exports WHERE session_id=?", (sid,))
        conn.execute("DELETE FROM sessions WHERE user_id=?", (user_id,))
        conn.execute("DELETE FROM users WHERE id=?", (user_id,))

# ══════════════════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;700&family=Noto+Sans+KR:wght@300;400;700&display=swap');
html,body,[class*="css"]{font-family:'Noto Sans KR',sans-serif}

.hdr{background:linear-gradient(135deg,#0f1117,#161920);border:1px solid #252830;
     border-radius:12px;padding:18px 24px;margin-bottom:16px;display:flex;align-items:center;gap:14px}
.hdr-title{font-family:'JetBrains Mono',monospace;font-size:20px;font-weight:700;
           background:linear-gradient(135deg,#00d4ff,#ff6b35);
           -webkit-background-clip:text;-webkit-text-fill-color:transparent;margin:0}
.hdr-sub{font-family:'JetBrains Mono',monospace;font-size:10px;color:#5a6072;margin:2px 0 0}
.ver-tag{margin-left:auto;font-family:'JetBrains Mono',monospace;font-size:10px;color:#00d4ff;
         border:1px solid #00d4ff;border-radius:4px;padding:2px 8px}

.card{background:#161920;border:1px solid #252830;border-radius:8px;padding:14px 16px;margin:4px 0}
.metric-val{font-family:'JetBrains Mono',monospace;font-size:26px;font-weight:700;color:#00d4ff}
.metric-lbl{font-size:11px;color:#5a6072;margin-top:2px}

.step{background:rgba(0,212,255,.1);border:1px solid rgba(0,212,255,.25);border-radius:5px;
      padding:5px 11px;font-family:'JetBrains Mono',monospace;font-size:10px;color:#00d4ff;
      letter-spacing:.1em;display:inline-block;margin-bottom:8px}

.info{background:rgba(0,212,255,.05);border:1px solid rgba(0,212,255,.18);
      border-radius:7px;padding:10px 14px;font-size:13px;color:#adb5c8;margin:6px 0}

.row-badge{display:inline-flex;align-items:center;gap:4px;padding:2px 8px;
           border-radius:20px;font-size:10px;font-weight:600}
.hi{background:rgba(0,229,160,.12);color:#00e5a0}
.md{background:rgba(255,209,102,.12);color:#ffd166}
.lo{background:rgba(255,71,87,.12);color:#ff4757}
.tag{background:rgba(0,212,255,.1);color:#00d4ff;padding:2px 7px;border-radius:4px;
     font-family:'JetBrains Mono',monospace;font-size:10px}

.tbl-wrap{border:1px solid #252830;border-radius:8px;overflow:hidden}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════
DEFAULTS = dict(
    results=[], video_path=None, video_name="", duration=0, fps=30,
    frame_preview=None, region=None, last_session_id=None,
    current_user_id=None, current_username=""
)
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ══════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════
def fmt_time(s: float) -> str:
    h, m = int(s//3600), int((s%3600)//60)
    ss, ms = int(s%60), int((s%1)*1000)
    return f"{h:02d}:{m:02d}:{ss:02d}" if h else f"{m:02d}:{ss:02d}.{ms:03d}"

def extract_frame(cap, t):
    cap.set(cv2.CAP_PROP_POS_MSEC, t*1000)
    ret, f = cap.read()
    return f if ret else None

def preprocess(img, mode):
    g = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    if mode == "inv": g = cv2.bitwise_not(g)
    if mode in ("auto","inv"):
        g = cv2.convertScaleAbs(g, alpha=1.8, beta=-30)
        g = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(4,4)).apply(g)
    return g

def run_ocr(gray, lang, mode, scale):
    h, w = gray.shape
    big = cv2.resize(gray, (w*scale, h*scale), interpolation=cv2.INTER_CUBIC)
    psm = "7" if (h*scale < 60 and mode != "all") else "6"
    cfg = f"--psm {psm} --oem 3"
    if mode == "dig":  cfg += " -c tessedit_char_whitelist=0123456789."
    elif mode == "num": cfg += " -c tessedit_char_whitelist=0123456789.-+eE/ "
    data = pytesseract.image_to_data(Image.fromarray(big), lang=lang, config=cfg,
                                     output_type=pytesseract.Output.DICT)
    words, confs = [], []
    for i, c in enumerate(data['conf']):
        try:
            ci = int(c)
        except Exception:
            continue
        if ci > 0:
            words.append(data['text'][i]); confs.append(ci)
    raw = ' '.join(words).strip()
    avg = int(sum(confs)/len(confs)) if confs else 0
    if mode == "dig":
        raw = '\n'.join(l.replace(' ','') for l in raw.split('\n') if any(c.isdigit() for c in l))
    elif mode == "num":
        raw = '  '.join(p for p in re.findall(r'[\d.+\-eE/]+', raw) if any(c.isdigit() for c in p))
    return raw.strip(), avg

def conf_cls(c): return "hi" if c>=70 else ("md" if c>=40 else "lo")
def conf_lbl(c): return ("🟢 높음" if c>=70 else ("🟡 보통" if c>=40 else "🔴 낮음"))
def fmt_bytes(b): return f"{b/1024:.1f} KB" if b < 1048576 else f"{b/1048576:.1f} MB"
def now_str(): return datetime.now().strftime("%Y%m%d_%H%M%S")

# ── Export builders ──────────────────────────────────────
def build_csv(results):
    df = pd.DataFrame(results)
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")

def build_xlsx(results, video_name=""):
    wb = Workbook()
    # Sheet1: Raw
    ws = wb.active; ws.title = "OCR 데이터"
    hdr_fill = PatternFill("solid", fgColor="00D4FF")
    hdr_font = Font(bold=True, color="000000", name="Arial", size=10)
    alt_fill = PatternFill("solid", fgColor="161920")
    headers = ["타임코드","시간(초)","신뢰도(%)","신뢰도 등급","추출값"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, r in enumerate(results, 2):
        ws.append([r["타임코드"], r["시간(초)"], r["신뢰도(%)"], r["신뢰도 등급"], r["추출값"]])
        if i % 2 == 0:
            for cell in ws[i]: cell.fill = alt_fill
    for col, w in zip(ws.columns, [16,12,12,14,50]):
        ws.column_dimensions[col[0].column_letter].width = w
    # Sheet2: Pivot
    ws2 = wb.create_sheet("숫자 피벗")
    ws2.append(["타임코드","시간(초)","값1","값2","값3"])
    for c in ws2[1]:
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
    for r in results:
        vals = [v.strip() for v in r["추출값"].replace('/',' ').split() if v.strip()]
        ws2.append([r["타임코드"], r["시간(초)"],
                    vals[0] if len(vals)>0 else "",
                    vals[1] if len(vals)>1 else "",
                    vals[2] if len(vals)>2 else ""])
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 16
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def build_pdf(results, video_name="", username=""):
    class PDF(FPDF):
        def header(self):
            self.set_fill_color(15,17,23)
            self.rect(0,0,297,210,'F')
            self.set_font("Helvetica","B",13)
            self.set_text_color(0,212,255)
            self.cell(0,8,"VIDEO OCR EXTRACTOR v3.0 - Results",ln=True)
            self.set_font("Helvetica","",8)
            self.set_text_color(90,96,114)
            info = f"User: {username}  |  Video: {video_name}  |  Rows: {len(results)}  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            self.cell(0,5,info,ln=True); self.ln(1)
        def footer(self):
            self.set_y(-11); self.set_font("Helvetica","",7)
            self.set_text_color(90,96,114)
            self.cell(0,8,f"Page {self.page_no()}",align="R")

    pdf = PDF(orientation="L",unit="mm",format="A4")
    pdf.add_page(); pdf.set_auto_page_break(True,12)
    cols  = ["타임코드","시간(초)","신뢰도(%)","추출값"]
    widths= [32,26,24,195]
    pdf.set_fill_color(0,212,255); pdf.set_text_color(0,0,0)
    pdf.set_font("Helvetica","B",8)
    for c,w in zip(cols,widths): pdf.cell(w,8,c,border=1,align="C",fill=True)
    pdf.ln()
    for i,r in enumerate(results):
        fill = i%2==0
        pdf.set_fill_color(22,25,32) if fill else pdf.set_fill_color(15,17,23)
        pdf.set_text_color(221,225,236); pdf.set_font("Helvetica","",7)
        for val,w in zip([r["타임코드"],str(r["시간(초)"]),f"{r['신뢰도(%)']}%",
                          r["추출값"].replace('\n',' / ')[:90]], widths):
            pdf.cell(w,6,val,border=1,fill=fill)
        pdf.ln()
    return pdf.output(dest="S").encode("latin-1")

# ══════════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════════
st.markdown("""
<div class="hdr">
  <div style="font-size:34px">🎬</div>
  <div>
    <p class="hdr-title">VIDEO OCR EXTRACTOR</p>
    <p class="hdr-sub">영역 선택 → 프레임 캡처 → 숫자 추출 → 이력 저장</p>
  </div>
  <div class="ver-tag">v3.0</div>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════
# SIDEBAR — 사용자 로그인 + 설정
# ══════════════════════════════════════════════════════════
with st.sidebar:
    # ── 사용자 선택 / 생성 ─────────────────
    st.markdown("### 👤 사용자")
    existing = [r["username"] for r in get_all_users()]
    mode_user = st.radio("", ["기존 사용자", "새 사용자 등록"], horizontal=True, label_visibility="collapsed")

    if mode_user == "새 사용자 등록":
        new_name = st.text_input("사용자명 입력", placeholder="예: 홍길동", key="new_user_input")
        if st.button("등록", use_container_width=True, type="primary"):
            if new_name.strip():
                uid = upsert_user(new_name.strip())
                st.session_state.current_user_id = uid
                st.session_state.current_username = new_name.strip()
                st.success(f"✓ '{new_name}' 등록/로그인 완료")
                st.rerun()
            else:
                st.error("사용자명을 입력하세요.")
    else:
        if existing:
            sel = st.selectbox("사용자 선택", existing, key="sel_user")
            if st.button("로그인", use_container_width=True, type="primary"):
                uid = upsert_user(sel)
                st.session_state.current_user_id = uid
                st.session_state.current_username = sel
                st.rerun()
        else:
            st.info("등록된 사용자가 없습니다.")

    # 현재 로그인 표시
    if st.session_state.current_username:
        st.markdown(f"""<div class="info">
            ✅ <b>{st.session_state.current_username}</b> 로그인 중
        </div>""", unsafe_allow_html=True)

    st.divider()

    # ── 비디오 업로드 ────────────────────────
    st.markdown("**📁 비디오 파일**")
    uploaded = st.file_uploader("MP4 · MOV · AVI · WebM",
                                type=["mp4","mov","avi","webm","mkv"],
                                label_visibility="collapsed")
    if uploaded:
        suffix = os.path.splitext(uploaded.name)[1]
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        tmp.write(uploaded.read()); tmp.flush()
        st.session_state.video_path = tmp.name
        st.session_state.video_name = uploaded.name
        cap = cv2.VideoCapture(tmp.name)
        st.session_state.fps      = cap.get(cv2.CAP_PROP_FPS) or 30
        n = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        st.session_state.duration = n / st.session_state.fps
        cap.release()
        st.success(f"✓ {uploaded.name}")
        st.caption(f"길이: {fmt_time(st.session_state.duration)}  |  FPS: {st.session_state.fps:.1f}")

    st.divider()

    # ── 추출 설정 ────────────────────────────
    st.markdown("**⚙️ 추출 설정**")
    c1, c2 = st.columns(2)
    with c1: interval = st.number_input("간격(초)", 0.1, 60.0, 1.0, 0.1, format="%.1f")
    with c2:
        lang_map = {"영어(숫자권장)":"eng","한국어+영어":"kor+eng","한국어":"kor"}
        lang = lang_map[st.selectbox("언어", list(lang_map.keys()))]
    c3, c4 = st.columns(2)
    with c3: t_start = st.number_input("시작(초)", 0.0, value=0.0, step=1.0)
    with c4:
        dur = st.session_state.duration
        t_end = st.number_input("종료(초)", 0.0, value=float(int(dur)) if dur>0 else 0.0, step=1.0, help="0=전체")
    c5, c6 = st.columns(2)
    with c5:
        mode_map = {"숫자 위주":"num","숫자만":"dig","일반 텍스트":"all"}
        mode = mode_map[st.selectbox("OCR 모드", list(mode_map.keys()))]
    with c6:
        scale = st.selectbox("확대", [3,2,4,1], format_func=lambda x: f"{x}×{'(권장)' if x==3 else ''}")
    prep_map = {"자동(그레이+대비)":"auto","반전+대비(밝은LCD)":"inv","원본 유지":"none"}
    prep = prep_map[st.selectbox("전처리", list(prep_map.keys()))]

    st.divider()
    # 미리보기 슬라이더
    st.markdown("**🎞️ 미리보기 시간**")
    if st.session_state.video_path and st.session_state.duration > 0:
        prev_sec = st.slider("", 0.0, float(int(st.session_state.duration)), 0.0, 1.0,
                             label_visibility="collapsed")
        cap = cv2.VideoCapture(st.session_state.video_path)
        f = extract_frame(cap, prev_sec); cap.release()
        if f is not None:
            st.session_state.frame_preview = cv2.cvtColor(f, cv2.COLOR_BGR2RGB)

# ══════════════════════════════════════════════════════════
# MAIN TABS
# ══════════════════════════════════════════════════════════
tab_ocr, tab_hist, tab_mgr = st.tabs(["🔬 OCR 추출", "📂 이력 / 파일 목록", "👥 사용자 관리"])

# ══════════════════════════════════════════════════════════
# TAB 1 — OCR 추출
# ══════════════════════════════════════════════════════════
with tab_ocr:
    if not st.session_state.current_username:
        st.warning("👈 사이드바에서 사용자를 선택하거나 등록하세요.")
        st.stop()

    left, right = st.columns([1.1, 1], gap="large")

    with left:
        st.markdown('<div class="step">STEP 1 — OCR 영역 선택</div>', unsafe_allow_html=True)
        if st.session_state.frame_preview is not None:
            fimg = st.session_state.frame_preview
            fh, fw = fimg.shape[:2]
            dw = 580; dh = int(fh*dw/fw)
            cr = st_canvas(
                fill_color="rgba(0,212,255,0.07)", stroke_width=2,
                stroke_color="#00d4ff",
                background_image=Image.fromarray(fimg),
                update_streamlit=True, width=dw, height=dh,
                drawing_mode="rect", key="canvas",
            )
            if cr.json_data and cr.json_data.get("objects"):
                obj = cr.json_data["objects"][-1]
                sx, sy = fw/dw, fh/dh
                rx,ry = int(obj.get("left",0)*sx), int(obj.get("top",0)*sy)
                rw,rh = int(obj.get("width",0)*sx), int(obj.get("height",0)*sy)
                if rw>10 and rh>10:
                    st.session_state.region = (rx,ry,rw,rh)

            r = st.session_state.region
            if r:
                c1,c2,c3,c4 = st.columns(4)
                for col,lbl,val in zip([c1,c2,c3,c4],["X","Y","W","H"],r):
                    with col:
                        st.markdown(f'<div class="card" style="text-align:center">'
                                    f'<div class="metric-val">{val}</div>'
                                    f'<div class="metric-lbl">{lbl}</div></div>',
                                    unsafe_allow_html=True)
                # Preview crop
                st.markdown("**🔍 영역 미리보기**")
                frame_bgr = cv2.cvtColor(fimg, cv2.COLOR_RGB2BGR)
                crop = frame_bgr[r[1]:r[1]+r[3], r[0]:r[0]+r[2]]
                if crop.size > 0:
                    g = preprocess(crop, prep)
                    bh,bw = g.shape
                    big = cv2.resize(g, (bw*int(scale), bh*int(scale)), interpolation=cv2.INTER_CUBIC)
                    st.image(big, use_column_width=True)
            else:
                st.markdown('<div class="info">↑ 캔버스에서 OCR 영역을 드래그하세요</div>',
                            unsafe_allow_html=True)
        else:
            st.markdown('<div class="info">← 비디오를 먼저 업로드하세요</div>', unsafe_allow_html=True)

    with right:
        st.markdown('<div class="step">STEP 2 — 추출 실행 및 저장</div>', unsafe_allow_html=True)

        can_run = st.session_state.video_path and st.session_state.duration > 0
        rc, clc = st.columns([2,1])
        with rc:
            run_btn = st.button("▶ OCR 추출 시작", disabled=not can_run,
                                use_container_width=True, type="primary")
        with clc:
            if st.button("🗑 초기화", use_container_width=True):
                st.session_state.results = []; st.session_state.last_session_id = None
                st.rerun()

        # ── OCR LOOP ──────────────────────────────────
        if run_btn and can_run:
            end_s = t_end if t_end > 0 else st.session_state.duration
            if t_start >= end_s:
                st.error("시작 ≥ 종료 시간")
            else:
                st.session_state.results = []
                ts_list = []
                t = t_start
                while t <= end_s + 1e-6:
                    ts_list.append(min(t, end_s))
                    t = round(t + interval, 6)

                total = len(ts_list)
                prog = st.progress(0, text="초기화...")
                stat = st.empty()
                live_tbl = st.empty()

                cap = cv2.VideoCapture(st.session_state.video_path)
                r = st.session_state.region

                for i, ts in enumerate(ts_list):
                    prog.progress(i/total, text=f"{i+1}/{total}  |  {fmt_time(ts)}")
                    stat.caption(f"처리: {fmt_time(ts)}")

                    frame = extract_frame(cap, ts)
                    if frame is None:
                        st.session_state.results.append({
                            "타임코드":fmt_time(ts),"시간(초)":round(ts,3),
                            "신뢰도(%)":0,"신뢰도 등급":"오류","추출값":"(프레임 실패)"})
                        continue

                    if r and r[2]>0 and r[3]>0:
                        fh,fw = frame.shape[:2]
                        x1,y1 = max(0,r[0]),max(0,r[1])
                        x2,y2 = min(fw,r[0]+r[2]),min(fh,r[1]+r[3])
                        crop = frame[y1:y2,x1:x2]
                    else:
                        crop = frame

                    if crop.size == 0:
                        st.session_state.results.append({
                            "타임코드":fmt_time(ts),"시간(초)":round(ts,3),
                            "신뢰도(%)":0,"신뢰도 등급":"오류","추출값":"(영역 오류)"})
                        continue

                    gray = preprocess(crop, prep)
                    ocr_lang = "eng" if mode in ("num","dig") else lang
                    text, conf = run_ocr(gray, ocr_lang, mode, int(scale))
                    st.session_state.results.append({
                        "타임코드":fmt_time(ts),"시간(초)":round(ts,3),
                        "신뢰도(%)":conf,"신뢰도 등급":conf_lbl(conf),
                        "추출값":text if text else "(없음)"})

                    if i % 5 == 0 and st.session_state.results:
                        live_tbl.dataframe(
                            pd.DataFrame(st.session_state.results[-15:])[["타임코드","신뢰도(%)","추출값"]],
                            use_container_width=True)

                cap.release()
                prog.progress(1.0, text=f"완료 — {len(st.session_state.results)}건 ✓")
                stat.empty(); live_tbl.empty()

                # ── DB 저장 ──────────────────────────
                sid = save_session(
                    user_id=st.session_state.current_user_id,
                    video_name=st.session_state.video_name,
                    interval_s=interval, ocr_mode=mode, lang=lang,
                    region=st.session_state.region,
                    results=st.session_state.results
                )
                st.session_state.last_session_id = sid
                st.success(f"✅ {len(st.session_state.results)}건 추출 완료 — 이력에 저장됨 (세션 #{sid})")

        # ── RESULTS + EXPORT ──────────────────────────
        st.markdown("---")
        st.markdown('<div class="step">STEP 3 — 결과 확인 및 내보내기</div>', unsafe_allow_html=True)
        results = st.session_state.results

        if results:
            df = pd.DataFrame(results)
            c1,c2,c3 = st.columns(3)
            with c1:
                st.markdown(f'<div class="card" style="text-align:center">'
                            f'<div class="metric-val">{len(df)}</div>'
                            f'<div class="metric-lbl">총 추출 건수</div></div>', unsafe_allow_html=True)
            with c2:
                avg = int(df["신뢰도(%)"].mean())
                st.markdown(f'<div class="card" style="text-align:center">'
                            f'<div class="metric-val">{avg}%</div>'
                            f'<div class="metric-lbl">평균 신뢰도</div></div>', unsafe_allow_html=True)
            with c3:
                hi = int((df["신뢰도(%)"]>=70).sum())
                st.markdown(f'<div class="card" style="text-align:center">'
                            f'<div class="metric-val">{hi}</div>'
                            f'<div class="metric-lbl">높음 (≥70%)</div></div>', unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            st.dataframe(df, use_container_width=True, height=300)

            st.markdown("**📦 내보내기**")
            d1,d2,d3,d4 = st.columns(4)
            sid = st.session_state.last_session_id
            uid = st.session_state.current_user_id
            vname = st.session_state.video_name
            uname = st.session_state.current_username
            ts_now = now_str()

            with d1:
                data = build_csv(results)
                fname = f"ocr_{ts_now}.csv"
                if st.download_button("📄 CSV", data=data, file_name=fname,
                                      mime="text/csv", use_container_width=True):
                    if sid: log_export(sid, uid, "CSV", fname, len(data))

            with d2:
                data = build_xlsx(results, vname)
                fname = f"ocr_{ts_now}.xlsx"
                if st.download_button("📊 Excel", data=data, file_name=fname,
                                      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                      use_container_width=True):
                    if sid: log_export(sid, uid, "Excel", fname, len(data))

            with d3:
                data = build_pdf(results, vname, uname)
                fname = f"ocr_{ts_now}.pdf"
                if st.download_button("📋 PDF", data=data, file_name=fname,
                                      mime="application/pdf", use_container_width=True):
                    if sid: log_export(sid, uid, "PDF", fname, len(data))

            with d4:
                tsv = "\n".join(f"[{r['타임코드']}]\t{r['추출값'].replace(chr(10),' ')}" for r in results)
                fname = f"ocr_{ts_now}.tsv"
                data = tsv.encode("utf-8-sig")
                if st.download_button("📎 TSV", data=data, file_name=fname,
                                      mime="text/tab-separated-values", use_container_width=True):
                    if sid: log_export(sid, uid, "TSV", fname, len(data))
        else:
            st.markdown('<div class="info" style="text-align:center;padding:40px">'
                        '<div style="font-size:40px;opacity:.12;margin-bottom:10px">📋</div>'
                        'OCR 추출 결과가 여기에 표시됩니다</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════
# TAB 2 — 이력 / 파일 목록
# ══════════════════════════════════════════════════════════
with tab_hist:
    st.markdown("### 📂 추출 이력 및 내보내기 파일 목록")

    uid_filter = st.session_state.current_user_id
    users = get_all_users()

    # 필터
    hcol1, hcol2 = st.columns([1,3])
    with hcol1:
        user_opts = ["전체 사용자"] + [u["username"] for u in users]
        sel_user_hist = st.selectbox("사용자 필터", user_opts, key="hist_user_filter")
        filter_uid = None
        if sel_user_hist != "전체 사용자":
            for u in users:
                if u["username"] == sel_user_hist:
                    filter_uid = u["id"]; break

    # ── 세션 이력 ────────────────────────────
    st.markdown("#### 📋 추출 세션 이력")
    sessions = get_sessions(filter_uid)

    if sessions:
        for s in sessions:
            with st.expander(
                f"#{s['id']}  {s['created']}  |  👤 {s['username']}  |  "
                f"🎬 {s['video_name'] or '(없음)'}  |  {s['total_rows']}건  |  신뢰도 {s['avg_conf']:.0f}%"
            ):
                cols = st.columns([2,1,1,1,1])
                cols[0].markdown(f"**모드:** `{s['ocr_mode']}`  **언어:** `{s['lang']}`  "
                                 f"**간격:** `{s['interval_s']}s`")
                if s['region']:
                    try:
                        rg = json.loads(s['region'])
                        cols[1].markdown(f"**영역:** `{rg}`")
                    except Exception:
                        pass

                rows = get_rows(s['id'])
                if rows:
                    df_s = pd.DataFrame([dict(r) for r in rows])
                    df_s = df_s[["timecode","time_sec","confidence","value"]]
                    df_s.columns = ["타임코드","시간(초)","신뢰도(%)","추출값"]
                    st.dataframe(df_s, use_container_width=True, height=220)

                    # Re-export buttons
                    ts_now = now_str()
                    results_for_export = [
                        {"타임코드":r["timecode"],"시간(초)":r["time_sec"],
                         "신뢰도(%)":r["confidence"],"신뢰도 등급":conf_lbl(r["confidence"]),
                         "추출값":r["value"]} for r in rows
                    ]
                    ex1,ex2,ex3 = st.columns(3)
                    with ex1:
                        st.download_button(f"📄 CSV 재다운로드",
                            data=build_csv(results_for_export),
                            file_name=f"ocr_s{s['id']}_{ts_now}.csv",
                            mime="text/csv", key=f"csv_{s['id']}", use_container_width=True)
                    with ex2:
                        st.download_button(f"📊 Excel 재다운로드",
                            data=build_xlsx(results_for_export, s["video_name"] or ""),
                            file_name=f"ocr_s{s['id']}_{ts_now}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"xlsx_{s['id']}", use_container_width=True)
                    with ex3:
                        st.download_button(f"📋 PDF 재다운로드",
                            data=build_pdf(results_for_export, s["video_name"] or "", s["username"]),
                            file_name=f"ocr_s{s['id']}_{ts_now}.pdf",
                            mime="application/pdf",
                            key=f"pdf_{s['id']}", use_container_width=True)

                if st.button(f"🗑 세션 #{s['id']} 삭제", key=f"del_s_{s['id']}"):
                    delete_session(s['id']); st.rerun()
    else:
        st.markdown('<div class="info">아직 저장된 이력이 없습니다.</div>', unsafe_allow_html=True)

    st.markdown("---")

    # ── 내보내기 파일 목록 ──────────────────
    st.markdown("#### 📁 내보내기 파일 이력")
    exports = get_exports(filter_uid)

    if exports:
        FMT_ICON = {"CSV":"📄","Excel":"📊","PDF":"📋","TSV":"📎"}
        rows_exp = []
        for e in exports:
            rows_exp.append({
                "번호": e["id"],
                "형식": FMT_ICON.get(e["fmt"],"📄") + " " + e["fmt"],
                "파일명": e["filename"],
                "크기": fmt_bytes(e["size_bytes"]) if e["size_bytes"] else "-",
                "사용자": e["username"],
                "비디오": e["video_name"] or "-",
                "생성일시": e["created"],
            })
        df_exp = pd.DataFrame(rows_exp)
        st.dataframe(df_exp, use_container_width=True, height=340)

        # Summary stats
        st.markdown("#### 📊 내보내기 통계")
        sc1,sc2,sc3,sc4 = st.columns(4)
        total_exp = len(exports)
        fmt_counts = {}
        for e in exports:
            fmt_counts[e["fmt"]] = fmt_counts.get(e["fmt"],0) + 1
        sc1.metric("총 내보내기", total_exp)
        sc2.metric("CSV", fmt_counts.get("CSV",0))
        sc3.metric("Excel", fmt_counts.get("Excel",0))
        sc4.metric("PDF", fmt_counts.get("PDF",0))
    else:
        st.markdown('<div class="info">아직 내보내기 파일 이력이 없습니다.</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════
# TAB 3 — 사용자 관리
# ══════════════════════════════════════════════════════════
with tab_mgr:
    st.markdown("### 👥 사용자 관리")
    users = get_all_users()

    if users:
        for u in users:
            sessions_u = get_sessions(u["id"])
            exports_u  = get_exports(u["id"])
            n_sess = len(sessions_u)
            n_exp  = len(exports_u)
            total_rows = sum(s["total_rows"] for s in sessions_u)

            with st.expander(f"👤 {u['username']}  |  세션 {n_sess}개  |  추출 {total_rows}건  |  내보내기 {n_exp}건"):
                i1,i2,i3,i4 = st.columns(4)
                i1.markdown(f'<div class="card" style="text-align:center">'
                            f'<div class="metric-val">{n_sess}</div>'
                            f'<div class="metric-lbl">세션</div></div>', unsafe_allow_html=True)
                i2.markdown(f'<div class="card" style="text-align:center">'
                            f'<div class="metric-val">{total_rows}</div>'
                            f'<div class="metric-lbl">총 추출 행</div></div>', unsafe_allow_html=True)
                i3.markdown(f'<div class="card" style="text-align:center">'
                            f'<div class="metric-val">{n_exp}</div>'
                            f'<div class="metric-lbl">내보내기</div></div>', unsafe_allow_html=True)
                i4.markdown(f'<div class="card" style="text-align:center">'
                            f'<div class="metric-val" style="font-size:13px">{u["created"][:10]}</div>'
                            f'<div class="metric-lbl">가입일</div></div>', unsafe_allow_html=True)

                # 최근 세션 5개
                if sessions_u:
                    st.markdown("**최근 세션**")
                    df_u = pd.DataFrame([{
                        "세션#": s["id"], "비디오": s["video_name"] or "-",
                        "추출건수": s["total_rows"], "평균신뢰도": f"{s['avg_conf']:.0f}%",
                        "생성일시": s["created"]
                    } for s in sessions_u[:5]])
                    st.dataframe(df_u, use_container_width=True, hide_index=True)

                # 내보내기 통계
                if exports_u:
                    fmt_cnt = {}
                    for e in exports_u:
                        fmt_cnt[e["fmt"]] = fmt_cnt.get(e["fmt"],0) + 1
                    st.markdown("**내보내기 현황:** " +
                        "  ".join(f'<span class="tag">{k}: {v}건</span>' for k,v in fmt_cnt.items()),
                        unsafe_allow_html=True)

                st.markdown("---")
                if st.button(f"⚠️ '{u['username']}' 및 모든 데이터 삭제",
                             key=f"del_u_{u['id']}", type="secondary"):
                    delete_user(u["id"])
                    if st.session_state.current_user_id == u["id"]:
                        st.session_state.current_user_id = None
                        st.session_state.current_username = ""
                    st.rerun()
    else:
        st.markdown('<div class="info">등록된 사용자가 없습니다. 사이드바에서 사용자를 등록하세요.</div>',
                    unsafe_allow_html=True)

    # ── 전체 통계 ────────────────────────────
    st.markdown("---")
    st.markdown("#### 📊 전체 통계")
    all_sessions = get_sessions()
    all_exports  = get_exports()
    all_rows_cnt = sum(s["total_rows"] for s in all_sessions)

    sc1,sc2,sc3,sc4 = st.columns(4)
    sc1.metric("전체 사용자", len(users))
    sc2.metric("전체 세션",   len(all_sessions))
    sc3.metric("전체 추출 행", all_rows_cnt)
    sc4.metric("전체 내보내기", len(all_exports))
